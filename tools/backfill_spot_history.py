# -*- coding: utf-8 -*-
"""一次性回填現貨歷史（Session Average 口徑）。

2026-08-05 建立。背景：8/04 發現主序列半年來記的是 High 欄不是均價，因 High÷均價
倍率在品項間會漂（當日 1.29~1.76），比值型訊號會被憑空造出漲跌，故整條作廢重來
（舊檔在 data/_archive/）。原判定「無法回填」是錯的——TrendForce 自家歷史圖表確實
要付費，但有兩個公開鏡像在每日爬同一張表：

  主源 hw-codekim/dram_price   xlsx  日頻 2025-08-19 起，欄位＝Session Average + Change
  補源 karakotaram/memory-spot-tracker  csv  2026-01-12 起，是唯一帶 daily_high 的免費日頻

兩者都在爬 dramexchange.com 首頁的 tb_NationalDramSpotPrice，跟我們每日抓的
TrendForce 頁是同一份資料（三個時間點逐格比對過，見下方 --verify）。

**這支只跑一次。** 之後靠 scripts/fetch_spot.py 每日累積。理由：那兩個 repo 都是
個人專案、無 license、單一維護者，長期依賴等於把資料源交給陌生人。

用法：
  python tools/backfill_spot_history.py --verify   # 只檢查不寫入
  python tools/backfill_spot_history.py            # 實際寫入
"""

import csv
import io
import json
import re
import sys
import urllib.request
from pathlib import Path

import openpyxl

DATA = Path(__file__).parent.parent / "data"
SPOT = DATA / "spot_history.json"

XLSX_URL = "https://raw.githubusercontent.com/hw-codekim/dram_price/main/DRAMeXchange_Semi_price.xlsx"
CSV_URL = "https://raw.githubusercontent.com/karakotaram/memory-spot-tracker/main/data/dram_spot.csv"
UA = {"User-Agent": "Mozilla/5.0 (memory-cycle-dashboard backfill)"}

TARGET = [
    "DDR5 16Gb (2Gx8) 4800/5600",
    "DDR4 16Gb (2Gx8) 3200",
    "DDR4 8Gb (1Gx8) 3200",
]
# 兩源均價的容許差。實測 264 組重疊有 28 組不同，最大 1.29%、方向隨機，
# 是兩個爬蟲抓到同一天不同盤中快照（TrendForce 一天更新兩次：14:40 與 18:10）。
# 超過容許差時不採用該日的 high，避免把 A 快照的均價配上 B 快照的高點。
TOL = 0.02


def normalize(name):
    """324 天裡品項名稱漂過兩種寫法，不統一會讓序列斷成兩截。

    DDR5 16G (2Gx8) …  →  DDR5 16Gb (2Gx8) …   （188 + 136 天）
    DDR4 16Gb (2Gx8)3200 →  DDR4 16Gb (2Gx8) 3200 （188 + 109 天）
    """
    s = re.sub(r"\b16G\b", "16Gb", str(name).strip())
    s = re.sub(r"\s+", " ", s)
    return re.sub(r"\)(\d)", r") \1", s)


def download(url):
    return urllib.request.urlopen(
        urllib.request.Request(url, headers=UA), timeout=90
    ).read()


def load_xlsx():
    """{品項: {日期: (session_average, session_change)}}"""
    wb = openpyxl.load_workbook(
        io.BytesIO(download(XLSX_URL)), read_only=True, data_only=True
    )
    out = {k: {} for k in TARGET}
    for row in list(wb[wb.sheetnames[0]].iter_rows(values_only=True))[1:]:
        if not row or row[0] is None:
            continue
        item = normalize(row[1])
        if item not in out or row[2] is None:
            continue
        d = str(row[0])[:10]
        chg = float(row[3]) if row[3] is not None else None
        out[item][d] = (float(row[2]), chg)
    return out


def load_csv():
    """{品項: {日期: (session_avg, daily_high)}}"""
    out = {k: {} for k in TARGET}
    for r in csv.DictReader(io.StringIO(download(CSV_URL).decode("utf-8-sig"))):
        item = normalize(r["product"])
        if item in out:
            out[item][r["date"]] = (float(r["session_avg"]), float(r["daily_high"]))
    return out


def build_series(xl, cs):
    series, stats = {}, {}
    for key in TARGET:
        pts, matched, mismatched = [], 0, 0
        for d in sorted(xl[key]):
            avg, chg = xl[key][d]
            pt = {"date": d, "price": avg}
            hit = cs[key].get(d)
            if hit:
                if abs(hit[0] - avg) <= TOL:
                    pt["price_high"] = hit[1]
                    matched += 1
                else:
                    mismatched += 1
            if chg is not None:
                pt["change_pct"] = chg
            pts.append(pt)
        series[key] = pts
        stats[key] = {
            "days": len(pts),
            "with_high": matched,
            "high_dropped_snapshot_mismatch": mismatched,
            "range": f"{pts[0]['date']} → {pts[-1]['date']}" if pts else "",
        }
    return series, stats


def merge_existing(series, existing):
    """把現有檔接在回填之後。

    現有檔唯一那批點是 2026-08-05 00:17 手動重置時抓的，抓到的其實是 08-04 的盤
    （TrendForce 現貨盤下午才更新，跨午夜跑到的是前一日）。回填最後一天也是 08-04
    且數值相同，直接合併會多出一個假的持平日，故對同值者只留回填那筆。
    """
    notes = []
    for key, entries in existing.get("series", {}).items():
        target = series.setdefault(key, [])
        have = {p["date"] for p in target}
        for pt in entries:
            if pt["date"] == "2026-08-05":
                # 同一次跨午夜執行留下的，包含無回填來源的 GDDR6：一律改標為實際盤期
                pt = dict(pt, date="2026-08-04")
            last = target[-1] if target else None
            if last and abs(last["price"] - pt["price"]) < 1e-6:
                notes.append(
                    f"  丟棄 {key} {pt['date']}：與回填 {last['date']} 同值＝跨午夜重複記錄"
                )
                continue
            if pt["date"] in have:
                notes.append(f"  丟棄 {key} {pt['date']}：日期與回填重複")
                continue
            target.append(pt)
        target.sort(key=lambda p: p["date"])
    return notes


def main():
    verify_only = "--verify" in sys.argv

    print("下載回填來源…")
    xl, cs = load_xlsx(), load_csv()
    series, stats = build_series(xl, cs)

    print("\n=== 回填結果 ===")
    for k, s in stats.items():
        print(
            f"  {k:<32} {s['days']:>4} 天  {s['range']}"
            f"  含高點 {s['with_high']}（快照不符捨棄 {s['high_dropped_snapshot_mismatch']}）"
        )

    existing = (
        json.loads(SPOT.read_text(encoding="utf-8-sig"))
        if SPOT.exists()
        else {"series": {}}
    )
    notes = merge_existing(series, existing)
    print("\n=== 與現有檔合併 ===")
    print("\n".join(notes) if notes else "  無需處理")
    for k in series:
        if k not in TARGET:
            print(f"  {k}：無回填來源，維持原樣（{len(series[k])} 點）")

    if verify_only:
        print("\n--verify：未寫入任何檔案。")
        return

    out = {
        "updated": existing.get("updated", ""),
        "series": series,
        "price_basis": "session_average",
        "restarted_at": "2026-08-05",
        "predecessor": existing.get("predecessor"),
        "backfill": {
            "date": "2026-08-05",
            "primary": XLSX_URL,
            "high_supplement": CSV_URL,
            "note": "兩源皆爬 dramexchange.com 首頁 tb_NationalDramSpotPrice，與每日抓的 TrendForce 頁同一份資料",
            "tolerance": TOL,
            "stats": stats,
        },
    }
    SPOT.write_text(json.dumps(out, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\n已寫入 {SPOT}")


if __name__ == "__main__":
    main()
